import io
import uuid as _uuid
from datetime import datetime, timezone

from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.dependencies import require_global_admin
from app.storage import save_logo
from app.models.club import Club
from app.models.league import League
from app.models.player import Player
from app.models.report import ScoutingReport
from app.models.user import RefreshToken, User
from app.schemas.admin import (
    AdminClubItem,
    AdminLeagueItem,
    AdminPlayerItem,
    AdminReportItem,
    AdminUserItem,
    AiAccessToggleRequest,
    BulkDeleteRequest,
    BulkDeleteResult,
    BulkImportResult,
    BulkImportRowError,
    CreateClubRequest,
    CreateLeagueRequest,
    CreatePlayerRequest,
    CreateReportRequest,
    CreateUserRequest,
    ListResponse,
    UpdateClubRequest,
    UpdateLeagueRequest,
    UpdatePlayerRequest,
    UpdateReportRequest,
    UpdateUserRequest,
)
from app.schemas.player import PlayerStats, UpdatePlayerStatsRequest
from app.security import hash_password
from app.utils.audit import record_audit
from app.utils.notifications import create_notification, format_role, notify_global_admins
from app.utils.telegram import send_report_notification
from app.utils.uuid import parse_uuid, parse_uuid_list

router = APIRouter(prefix="/admin", tags=["admin"])


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
_VALID_POSITIONS = {"GK", "CB", "LB", "RB", "CDM", "CM", "AM", "CAM", "LW", "RW", "CF", "ST"}

_LOGO_MAX_BYTES = 2 * 1024 * 1024
_LOGO_SIGNATURES: dict[bytes, str] = {
    b"\x89PNG\r\n\x1a\n": "png",
    b"\xff\xd8\xff": "jpg",
}
_WEBP_RIFF = b"RIFF"
_WEBP_MARKER = b"WEBP"


def _detect_image_ext(data: bytes) -> str | None:
    for sig, ext in _LOGO_SIGNATURES.items():
        if data[:len(sig)] == sig:
            return ext
    if data[:4] == _WEBP_RIFF and len(data) >= 12 and data[8:12] == _WEBP_MARKER:
        return "webp"
    return None


async def _read_validated_logo(file: UploadFile) -> tuple[str, bytes]:
    raw = await file.read()
    if len(raw) > _LOGO_MAX_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Logo must not exceed 2 MB.")
    ext = _detect_image_ext(raw)
    if not ext:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Only PNG, JPEG, and WebP images are accepted.",
        )
    return ext, raw


async def _load_import_rows(file: UploadFile) -> list[tuple]:
    import openpyxl

    if file.content_type not in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) and not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Only Excel files (.xlsx, .xls) are accepted.")

    raw = await file.read()
    if len(raw) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File exceeds the 5 MB limit.")

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Could not parse the Excel file. Make sure it is a valid .xlsx or .xls file.")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="The file has no data rows. Row 1 must be the header.")
    return rows


def _get_league_or_404(db: Session, league_id: str) -> League:
    lid = parse_uuid(league_id, "league_id format")
    league = db.query(League).filter(League.id == lid).first()
    if not league:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="League not found.")
    return league


def _get_user_or_404(db: Session, user_id: str) -> User:
    uid = parse_uuid(user_id, "user_id format")
    user = db.query(User).filter(User.id == uid, User.deleted_at.is_(None)).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found.")
    return user


def _get_club_or_404(db: Session, club_id: str) -> Club:
    cid = parse_uuid(club_id, "club_id format")
    club = db.query(Club).filter(Club.id == cid, Club.deleted_at.is_(None)).first()
    if not club:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Club not found.")
    return club


def _get_player_or_404(db: Session, player_id: str) -> Player:
    pid = parse_uuid(player_id, "player_id format")
    player = db.query(Player).filter(Player.id == pid).first()
    if not player:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Player not found.")
    return player


def _get_report_or_404(db: Session, report_id: str) -> ScoutingReport:
    rid = parse_uuid(report_id, "report_id format")
    report = db.query(ScoutingReport).filter(ScoutingReport.id == rid).first()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found.")
    return report


def _resolve_club(db: Session, club_id: str | None) -> tuple[_uuid.UUID | None, Club | None]:
    if not club_id:
        return None, None
    club_uuid = parse_uuid(club_id, "club_id format")
    club = db.query(Club).filter(Club.id == club_uuid, Club.deleted_at.is_(None)).first()
    if not club:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Club not found.")
    return club_uuid, club


def _resolve_league(db: Session, league_id: str | None) -> tuple[_uuid.UUID | None, League | None]:
    if not league_id:
        return None, None
    league_uuid = parse_uuid(league_id, "league_id format")
    league = db.query(League).filter(League.id == league_uuid).first()
    if not league:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="League not found.")
    return league_uuid, league


def _league_club_count(db: Session, league_id) -> int:
    return (
        db.query(func.count(Club.id))
        .filter(Club.league_id == league_id, Club.deleted_at.is_(None))
        .scalar()
    ) or 0


def _club_counts(db: Session, club_id) -> tuple[int, int]:
    scout_count = (
        db.query(func.count(User.id))
        .filter(User.club_id == club_id, User.role == "scout", User.deleted_at.is_(None))
        .scalar()
    ) or 0
    player_count = (
        db.query(func.count(Player.id))
        .filter(Player.club_id == club_id)
        .scalar()
    ) or 0
    return scout_count, player_count


def _league_item(league: League, club_count: int) -> AdminLeagueItem:
    return AdminLeagueItem(
        id=str(league.id),
        name=league.name,
        country=league.country,
        logo_url=league.logo_url,
        club_count=club_count,
        created_at=league.created_at,
    )


def _user_item(user: User, club_name: str | None) -> AdminUserItem:
    return AdminUserItem(
        id=str(user.id),
        email=user.email,
        first_name=user.first_name,
        last_name=user.last_name,
        role=user.role,
        club_id=str(user.club_id) if user.club_id else None,
        club_name=club_name,
        status=user.status,
        ai_access=user.ai_access,
        created_at=user.created_at,
    )


def _club_item(club: Club, league_name: str | None, scout_count: int, player_count: int) -> AdminClubItem:
    return AdminClubItem(
        id=str(club.id),
        name=club.name,
        country=club.country,
        league=league_name or "—",
        league_id=str(club.league_id) if club.league_id else None,
        logo_url=club.logo_url,
        scout_count=scout_count,
        player_count=player_count,
        status=club.status,
        created_at=club.created_at,
    )


def _player_item(player: Player, club_name: str | None) -> AdminPlayerItem:
    return AdminPlayerItem(
        id=str(player.id),
        first_name=player.first_name,
        last_name=player.last_name,
        date_of_birth=player.date_of_birth,
        nationality=player.nationality,
        position=player.position,
        club_name=club_name,
        market_value=player.market_value,
        status=player.status,
        stats=PlayerStats.from_player(player),
        created_at=player.created_at,
    )


def _report_item(report: ScoutingReport, scout_name: str) -> AdminReportItem:
    return AdminReportItem(
        id=str(report.id),
        player_id=str(report.player_id) if report.player_id else None,
        player_name=report.player_name,
        position=report.position,
        scout_name=scout_name,
        rating=report.rating,
        status=report.status,
        notes=report.notes,
        created_at=report.created_at,
    )


# ---------------------------------------------------------------------------
# Leagues
# ---------------------------------------------------------------------------

@router.get("/leagues", response_model=ListResponse[AdminLeagueItem])
def list_leagues(
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    leagues = db.query(League).order_by(League.name).all()
    club_counts = {
        row.league_id: row.cnt
        for row in db.query(Club.league_id, func.count(Club.id).label("cnt"))
        .filter(Club.league_id.isnot(None), Club.deleted_at.is_(None))
        .group_by(Club.league_id)
        .all()
    }
    items = [_league_item(l, club_counts.get(l.id, 0)) for l in leagues]
    return ListResponse(items=items, total=len(items))


@router.post("/leagues", response_model=AdminLeagueItem, status_code=status.HTTP_201_CREATED)
def create_league(
    body: CreateLeagueRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    new_league = League(
        name=body.name.strip(),
        country=body.country.strip(),
    )
    db.add(new_league)
    db.commit()
    db.refresh(new_league)
    return _league_item(new_league, 0)


@router.post("/leagues/bulk-delete", response_model=BulkDeleteResult)
def bulk_delete_leagues(
    body: BulkDeleteRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    uuids = parse_uuid_list(body.ids)
    if not uuids:
        return BulkDeleteResult(deleted=0)
    count = db.query(League).filter(League.id.in_(uuids)).delete(synchronize_session=False)
    db.commit()
    return BulkDeleteResult(deleted=count)


@router.put("/leagues/{league_id}", response_model=AdminLeagueItem)
def update_league(
    league_id: str,
    body: UpdateLeagueRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    league = _get_league_or_404(db, league_id)

    league.name = body.name.strip()
    league.country = body.country.strip()
    db.commit()
    db.refresh(league)

    return _league_item(league, _league_club_count(db, league.id))


@router.delete("/leagues/{league_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_league(
    league_id: str,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
) -> None:
    league = _get_league_or_404(db, league_id)
    db.delete(league)
    db.commit()


@router.post("/leagues/{league_id}/logo", response_model=AdminLeagueItem)
async def upload_league_logo(
    league_id: str,
    file: UploadFile = File(...),
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    league = _get_league_or_404(db, league_id)
    ext, raw = await _read_validated_logo(file)

    league.logo_url = save_logo(f"league_{league_id}", ext, raw)
    db.commit()
    db.refresh(league)

    return _league_item(league, _league_club_count(db, league.id))


@router.get("/scouts")
def list_scouts(
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    scouts = (
        db.query(User)
        .filter(User.role == "scout", User.deleted_at.is_(None))
        .order_by(User.first_name, User.last_name)
        .all()
    )
    return [{"id": str(s.id), "name": f"{s.first_name} {s.last_name}"} for s in scouts]


# ---------------------------------------------------------------------------
# Users
# ---------------------------------------------------------------------------

@router.get("/users", response_model=ListResponse[AdminUserItem])
def list_users(
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(User, Club.name.label("club_name"))
        .outerjoin(Club, User.club_id == Club.id)
        .filter(User.deleted_at.is_(None))
        .order_by(User.created_at.desc())
        .all()
    )
    items = [_user_item(user, club_name) for user, club_name in rows]
    return ListResponse(items=items, total=len(items))


@router.post("/users", response_model=AdminUserItem, status_code=status.HTTP_201_CREATED)
def create_user(
    body: CreateUserRequest,
    request: Request,
    admin: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    club_uuid, club = _resolve_club(db, body.club_id)

    if body.role == "player" and not body.position:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Position is required when creating a player account.",
        )

    new_user = User(
        email=body.email.lower(),
        password_hash=hash_password(body.password),
        first_name=body.first_name.strip(),
        last_name=body.last_name.strip(),
        role=body.role,
        club_id=club_uuid,
        status=body.status,
    )

    try:
        db.add(new_user)
        db.flush()
        if body.role == "player":
            db.add(Player(
                user_id=new_user.id,
                first_name=body.first_name.strip(),
                last_name=body.last_name.strip(),
                position=body.position.strip().upper(),
                club_id=club_uuid,
                availability_status="under_contract" if club_uuid else "free_agent",
                status="active",
            ))

        notify_global_admins(
            db,
            "profile",
            "New User Created",
            f"{body.first_name.strip()} {body.last_name.strip()} was created as {format_role(body.role)}.",
        )

        record_audit(
            db, "user.create", actor=admin, target_type="user", target_id=new_user.id,
            request=request, detail=f"Created {new_user.email} as {body.role}.",
        )

        db.commit()
        db.refresh(new_user)
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A user with this email already exists.",
        )

    return _user_item(new_user, club.name if club else None)


@router.post("/users/{user_id}/ai-access", response_model=AdminUserItem)
def set_ai_access(
    user_id: str,
    body: AiAccessToggleRequest,
    request: Request,
    admin: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    user = _get_user_or_404(db, user_id)
    if user.role != "scout":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="AI access can only be granted to scouts.",
        )

    if user.ai_access != body.enabled:
        user.ai_access = body.enabled
        user.updated_at = datetime.now(timezone.utc)
        record_audit(
            db,
            "ai_access.grant" if body.enabled else "ai_access.revoke",
            actor=admin,
            target_type="user",
            target_id=user.id,
            request=request,
            detail=f"AI access {'granted' if body.enabled else 'revoked'} for {user.email}.",
        )
        create_notification(
            db,
            user.id,
            "profile",
            "AI Assistant Access " + ("Granted" if body.enabled else "Revoked"),
            "You now have access to the AI Assistant."
            if body.enabled
            else "Your access to the AI Assistant has been revoked.",
        )
        db.commit()
        db.refresh(user)

    club_name = None
    if user.club_id:
        club = db.query(Club).filter(Club.id == user.club_id).first()
        club_name = club.name if club else None

    return _user_item(user, club_name)


# ---------------------------------------------------------------------------
# Clubs
# ---------------------------------------------------------------------------

@router.get("/clubs", response_model=ListResponse[AdminClubItem])
def list_clubs(
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Club, League.name.label("league_name"))
        .outerjoin(League, Club.league_id == League.id)
        .filter(Club.deleted_at.is_(None))
        .order_by(Club.created_at.desc())
        .all()
    )

    scout_counts = {
        row.club_id: row.cnt
        for row in db.query(User.club_id, func.count(User.id).label("cnt"))
        .filter(User.role == "scout", User.club_id.isnot(None), User.deleted_at.is_(None))
        .group_by(User.club_id)
        .all()
    }

    player_counts = {
        row.club_id: row.cnt
        for row in db.query(Player.club_id, func.count(Player.id).label("cnt"))
        .filter(Player.club_id.isnot(None))
        .group_by(Player.club_id)
        .all()
    }

    items = [
        _club_item(club, league_name, scout_counts.get(club.id, 0), player_counts.get(club.id, 0))
        for club, league_name in rows
    ]
    return ListResponse(items=items, total=len(items))


@router.post("/clubs", response_model=AdminClubItem, status_code=status.HTTP_201_CREATED)
def create_club(
    body: CreateClubRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    league_uuid, league = _resolve_league(db, body.league_id)

    new_club = Club(
        name=body.name.strip(),
        short_name=body.short_name.strip() if body.short_name else None,
        country=body.country.strip(),
        league_id=league_uuid,
        status=body.status,
    )

    db.add(new_club)
    db.commit()
    db.refresh(new_club)

    return _club_item(new_club, league.name if league else None, 0, 0)


# ---------------------------------------------------------------------------
# Users – Bulk delete
# ---------------------------------------------------------------------------

@router.post("/users/bulk-delete", response_model=BulkDeleteResult)
def bulk_delete_users(
    body: BulkDeleteRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    uuids = parse_uuid_list(body.ids)
    if not uuids:
        return BulkDeleteResult(deleted=0)
    now = datetime.now(timezone.utc)
    count = (
        db.query(User)
        .filter(User.id.in_(uuids), User.deleted_at.is_(None))
        .update({User.deleted_at: now}, synchronize_session=False)
    )
    db.commit()
    return BulkDeleteResult(deleted=count)


# ---------------------------------------------------------------------------
# Clubs – Bulk import
# ---------------------------------------------------------------------------

@router.get("/clubs/bulk-import/template")
def download_clubs_template(_: User = Depends(require_global_admin)):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clubs"
    headers = ["name", "country", "league", "status"]
    ws.append(headers)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clubs_template.xlsx"},
    )


@router.post("/clubs/bulk-import", response_model=BulkImportResult)
async def bulk_import_clubs(
    file: UploadFile = File(...),
    current_admin: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    rows = await _load_import_rows(file)

    all_leagues = db.query(League).all()
    leagues_by_name = {l.name.strip().lower(): l for l in all_leagues}
    leagues_by_country: dict[str, list[str]] = {}
    for l in all_leagues:
        key = l.country.strip().lower()
        leagues_by_country.setdefault(key, []).append(l.name)

    _VALID_CLUB_STATUSES = {"active", "pending", "suspended"}

    errors: list[BulkImportRowError] = []
    created = 0

    for i, row in enumerate(rows, start=2):
        name = str(row[0]).strip() if row[0] is not None else ""
        country = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        league_raw = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        status_raw = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else "active"

        row_errors: list[BulkImportRowError] = []

        if not name:
            row_errors.append(BulkImportRowError(row=i, field="name", message="Club name is required."))
        elif len(name) > 200:
            row_errors.append(BulkImportRowError(row=i, field="name", message=f"Club name exceeds 200 characters (got {len(name)})."))

        if not country:
            row_errors.append(BulkImportRowError(row=i, field="country", message="Country is required."))
        elif len(country) > 100:
            row_errors.append(BulkImportRowError(row=i, field="country", message=f"Country exceeds 100 characters (got {len(country)})."))

        league_obj = None
        if not league_raw:
            available = sorted(leagues_by_country.get(country.strip().lower(), [])) if country else []
            if available:
                row_errors.append(BulkImportRowError(row=i, field="league", message=f"League is required. Available leagues for '{country}': {', '.join(available)}."))
            else:
                row_errors.append(BulkImportRowError(row=i, field="league", message="League is required. Use an existing league name."))
        else:
            league_obj = leagues_by_name.get(league_raw.lower())
            if not league_obj:
                row_errors.append(BulkImportRowError(row=i, field="league", message=f"League '{league_raw}' not found. Use an existing league name."))
            elif country and league_obj.country.strip().lower() != country.strip().lower():
                available = sorted(leagues_by_country.get(country.strip().lower(), []))
                suggestion = f" Available leagues for '{country}': {', '.join(available)}." if available else ""
                row_errors.append(BulkImportRowError(row=i, field="league", message=f"League '{league_raw}' belongs to {league_obj.country}, not '{country}'.{suggestion}"))
                league_obj = None

        if status_raw not in _VALID_CLUB_STATUSES:
            row_errors.append(BulkImportRowError(row=i, field="status", message=f"Invalid status '{status_raw}'. Must be one of: active, pending, suspended."))

        if row_errors:
            errors.extend(row_errors)
            continue

        new_club = Club(
            name=name,
            country=country,
            league_id=league_obj.id if league_obj else None,
            status=status_raw,
        )
        db.add(new_club)
        created += 1

    if created:
        notify_global_admins(
            db,
            "profile",
            "Bulk Import Successful",
            f"{created} club{'s' if created != 1 else ''} imported successfully.",
        )
        db.commit()

    return BulkImportResult(created=created, errors=errors)


# ---------------------------------------------------------------------------
# Players
# ---------------------------------------------------------------------------

@router.get("/players", response_model=ListResponse[AdminPlayerItem])
def list_players(
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Player, Club.name.label("club_name"))
        .outerjoin(Club, Player.club_id == Club.id)
        .order_by(Player.created_at.desc())
        .all()
    )
    items = [_player_item(player, club_name) for player, club_name in rows]
    return ListResponse(items=items, total=len(items))


@router.post("/players", response_model=AdminPlayerItem, status_code=status.HTTP_201_CREATED)
def create_player(
    body: CreatePlayerRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    club_uuid, club = _resolve_club(db, body.club_id)

    new_player = Player(
        first_name=body.first_name.strip(),
        last_name=body.last_name.strip(),
        position=body.position.strip(),
        nationality=body.nationality.strip() if body.nationality else None,
        date_of_birth=body.date_of_birth,
        club_id=club_uuid,
        availability_status="under_contract" if club_uuid else "free_agent",
        market_value=body.market_value,
        status=body.status,
    )

    db.add(new_player)
    db.commit()
    db.refresh(new_player)

    return _player_item(new_player, club.name if club else None)


# ---------------------------------------------------------------------------
# Clubs – Bulk delete
# ---------------------------------------------------------------------------

@router.post("/clubs/bulk-delete", response_model=BulkDeleteResult)
def bulk_delete_clubs(
    body: BulkDeleteRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    uuids = parse_uuid_list(body.ids)
    if not uuids:
        return BulkDeleteResult(deleted=0)
    now = datetime.now(timezone.utc)
    count = (
        db.query(Club)
        .filter(Club.id.in_(uuids), Club.deleted_at.is_(None))
        .update({Club.deleted_at: now}, synchronize_session=False)
    )
    db.commit()
    return BulkDeleteResult(deleted=count)


# ---------------------------------------------------------------------------
# Players – Bulk import
# ---------------------------------------------------------------------------

@router.get("/players/bulk-import/template")
def download_players_template(_: User = Depends(require_global_admin)):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Players"
    headers = ["first_name", "last_name", "position", "nationality", "date_of_birth", "club_name", "status", "market_value_millions"]
    ws.append(headers)
    ws.append(["Luka", "Modric", "CM", "Croatian", "1985-09-09", "Real Madrid", "active", 5])
    ws.append(["Erling", "Haaland", "ST", "Norwegian", "2000-07-21", "Manchester City", "active", 180])
    col_widths = [15, 15, 12, 14, 16, 20, 10, 22]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=players_template.xlsx"},
    )


@router.post("/players/bulk-import", response_model=BulkImportResult)
async def bulk_import_players(
    file: UploadFile = File(...),
    current_admin: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    from datetime import date as _date

    rows = await _load_import_rows(file)

    clubs_by_name = {c.name.strip().lower(): c for c in db.query(Club).filter(Club.deleted_at.is_(None)).all()}
    _VALID_PLAYER_STATUSES = {"active", "injured"}

    errors: list[BulkImportRowError] = []
    created = 0

    for i, row in enumerate(rows, start=2):
        first_name = str(row[0]).strip() if row[0] is not None else ""
        last_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        position = str(row[2]).strip().upper() if len(row) > 2 and row[2] is not None else ""
        nationality = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        dob_raw = row[4] if len(row) > 4 else None
        club_raw = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        status_raw = str(row[6]).strip().lower() if len(row) > 6 and row[6] is not None else "active"
        market_val_raw = row[7] if len(row) > 7 else None

        row_errors: list[BulkImportRowError] = []

        if not first_name:
            row_errors.append(BulkImportRowError(row=i, field="first_name", message="First name is required."))
        elif len(first_name) > 100:
            row_errors.append(BulkImportRowError(row=i, field="first_name", message=f"First name exceeds 100 characters (got {len(first_name)})."))

        if not last_name:
            row_errors.append(BulkImportRowError(row=i, field="last_name", message="Last name is required."))
        elif len(last_name) > 100:
            row_errors.append(BulkImportRowError(row=i, field="last_name", message=f"Last name exceeds 100 characters (got {len(last_name)})."))

        if not position:
            row_errors.append(BulkImportRowError(row=i, field="position", message="Position is required."))
        elif position not in _VALID_POSITIONS:
            row_errors.append(BulkImportRowError(row=i, field="position", message=f"Invalid position '{position}'. Must be one of: {', '.join(sorted(_VALID_POSITIONS))}."))

        if nationality and len(nationality) > 100:
            row_errors.append(BulkImportRowError(row=i, field="nationality", message=f"Nationality exceeds 100 characters (got {len(nationality)})."))

        dob: _date | None = None
        if dob_raw is not None and str(dob_raw).strip():
            if isinstance(dob_raw, _date):
                dob = dob_raw
            else:
                try:
                    dob = _date.fromisoformat(str(dob_raw).strip())
                except ValueError:
                    row_errors.append(BulkImportRowError(row=i, field="date_of_birth", message=f"Invalid date format '{dob_raw}'. Use YYYY-MM-DD."))

        club_obj = None
        if club_raw:
            club_obj = clubs_by_name.get(club_raw.lower())
            if not club_obj:
                row_errors.append(BulkImportRowError(row=i, field="club_name", message=f"Club '{club_raw}' not found. Leave blank or use an existing club name."))

        if status_raw not in _VALID_PLAYER_STATUSES:
            row_errors.append(BulkImportRowError(row=i, field="status", message=f"Invalid status '{status_raw}'. Must be one of: active, injured."))

        market_value: int | None = None
        if market_val_raw is not None and str(market_val_raw).strip():
            try:
                mv_millions = float(str(market_val_raw))
                if mv_millions < 0:
                    raise ValueError
                market_value = int(mv_millions * 1_000_000)
            except (ValueError, TypeError):
                row_errors.append(BulkImportRowError(row=i, field="market_value_millions", message=f"Invalid market value '{market_val_raw}'. Enter a non-negative number in millions (e.g. 50 = €50M)."))

        if row_errors:
            errors.extend(row_errors)
            continue

        new_player = Player(
            first_name=first_name,
            last_name=last_name,
            position=position,
            nationality=nationality or None,
            date_of_birth=dob,
            club_id=club_obj.id if club_obj else None,
            availability_status="under_contract" if club_obj else "free_agent",
            market_value=market_value,
            status=status_raw,
        )
        db.add(new_player)
        created += 1

    if created:
        notify_global_admins(
            db,
            "profile",
            "Bulk Import Successful",
            f"{created} player{'s' if created != 1 else ''} imported successfully.",
        )
        db.commit()

    return BulkImportResult(created=created, errors=errors)


# ---------------------------------------------------------------------------
# Players – Bulk delete
# ---------------------------------------------------------------------------

@router.post("/players/bulk-delete", response_model=BulkDeleteResult)
def bulk_delete_players(
    body: BulkDeleteRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    uuids = parse_uuid_list(body.ids)
    if not uuids:
        return BulkDeleteResult(deleted=0)
    count = db.query(Player).filter(Player.id.in_(uuids)).delete(synchronize_session=False)
    db.commit()
    return BulkDeleteResult(deleted=count)


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

@router.get("/reports", response_model=ListResponse[AdminReportItem])
def list_reports(
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(ScoutingReport, User.first_name, User.last_name)
        .join(User, ScoutingReport.scout_id == User.id)
        .order_by(ScoutingReport.created_at.desc())
        .all()
    )
    items = [
        _report_item(report, f"{first_name} {last_name}")
        for report, first_name, last_name in rows
    ]
    return ListResponse(items=items, total=len(items))


@router.post("/reports", response_model=AdminReportItem, status_code=status.HTTP_201_CREATED)
def create_report(
    body: CreateReportRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    scout_uuid = parse_uuid(body.scout_id, "scout_id format")

    scout = (
        db.query(User)
        .filter(User.id == scout_uuid, User.role == "scout", User.deleted_at.is_(None))
        .first()
    )
    if not scout:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Scout not found.",
        )

    player_uuid = None
    if body.player_id:
        player_uuid = _get_player_or_404(db, body.player_id).id

    new_report = ScoutingReport(
        scout_id=scout_uuid,
        player_id=player_uuid,
        player_name=body.player_name.strip(),
        position=body.position.strip(),
        rating=body.rating,
        status=body.status,
        notes=body.notes,
    )

    db.add(new_report)
    db.commit()
    db.refresh(new_report)

    return _report_item(new_report, f"{scout.first_name} {scout.last_name}")


# ---------------------------------------------------------------------------
# Users – Update & Delete
# ---------------------------------------------------------------------------

@router.put("/users/{user_id}", response_model=AdminUserItem)
def update_user(
    user_id: str,
    body: UpdateUserRequest,
    request: Request,
    admin: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    user = _get_user_or_404(db, user_id)

    old_role = user.role
    old_status = user.status

    new_email = body.email.lower()
    if user.email != new_email:
        conflict = (
            db.query(User)
            .filter(User.email == new_email, User.deleted_at.is_(None), User.id != user.id)
            .first()
        )
        if conflict:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A user with this email already exists.")

    club_uuid, club = _resolve_club(db, body.club_id)

    user.email = new_email
    user.first_name = body.first_name.strip()
    user.last_name = body.last_name.strip()
    user.role = body.role
    user.club_id = club_uuid
    user.status = body.status
    user.updated_at = datetime.now(timezone.utc)

    if body.role != "scout" and user.ai_access:
        user.ai_access = False

    if body.role != old_role or body.status != old_status:
        record_audit(
            db, "user.update", actor=admin, target_type="user", target_id=user.id,
            request=request,
            detail=f"role {old_role}->{body.role}, status {old_status}->{body.status}.",
        )

    player = db.query(Player).filter(Player.user_id == user.id).first()
    if player:
        player.club_id = club_uuid
        player.first_name = body.first_name.strip()
        player.last_name = body.last_name.strip()
        player.availability_status = "under_contract" if club_uuid else "free_agent"

    db.commit()
    db.refresh(user)

    return _user_item(user, club.name if club else None)


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: str,
    request: Request,
    admin: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
) -> None:
    user = _get_user_or_404(db, user_id)

    user.deleted_at = datetime.now(timezone.utc)
    db.query(RefreshToken).filter(
        RefreshToken.user_id == user.id,
        RefreshToken.revoked.is_(False),
    ).update({"revoked": True}, synchronize_session=False)
    record_audit(
        db, "user.delete", actor=admin, target_type="user", target_id=user.id,
        request=request, detail=f"Soft-deleted {user.email}; sessions revoked.",
    )
    db.commit()


# ---------------------------------------------------------------------------
# Clubs – Update & Delete
# ---------------------------------------------------------------------------

@router.put("/clubs/{club_id}", response_model=AdminClubItem)
def update_club(
    club_id: str,
    body: UpdateClubRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    club = _get_club_or_404(db, club_id)
    league_uuid, league = _resolve_league(db, body.league_id)

    club.name = body.name.strip()
    club.country = body.country.strip()
    club.league_id = league_uuid
    club.status = body.status
    club.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(club)

    scout_count, player_count = _club_counts(db, club.id)
    return _club_item(club, league.name if league else None, scout_count, player_count)


@router.delete("/clubs/{club_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_club(
    club_id: str,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
) -> None:
    club = _get_club_or_404(db, club_id)

    club.deleted_at = datetime.now(timezone.utc)
    db.commit()


# ---------------------------------------------------------------------------
# Clubs – Logo upload
# ---------------------------------------------------------------------------

@router.post("/clubs/{club_id}/logo", response_model=AdminClubItem)
async def upload_club_logo(
    club_id: str,
    file: UploadFile = File(...),
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    club = _get_club_or_404(db, club_id)
    ext, raw = await _read_validated_logo(file)

    club.logo_url = save_logo(club_id, ext, raw)
    club.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(club)

    league_name = None
    if club.league_id:
        league = db.query(League).filter(League.id == club.league_id).first()
        if league:
            league_name = league.name

    scout_count, player_count = _club_counts(db, club.id)
    return _club_item(club, league_name, scout_count, player_count)


# ---------------------------------------------------------------------------
# Players – Update & Delete
# ---------------------------------------------------------------------------

@router.put("/players/{player_id}", response_model=AdminPlayerItem)
def update_player(
    player_id: str,
    body: UpdatePlayerRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    player = _get_player_or_404(db, player_id)
    club_uuid, club = _resolve_club(db, body.club_id)

    old_status = player.status
    player.first_name = body.first_name.strip()
    player.last_name = body.last_name.strip()
    player.position = body.position.strip()
    player.nationality = body.nationality.strip() if body.nationality else None
    player.date_of_birth = body.date_of_birth
    player.club_id = club_uuid
    player.availability_status = "under_contract" if club_uuid else "free_agent"
    player.market_value = body.market_value
    player.status = body.status

    if body.status != old_status:
        from app.models.saved_prospect import SavedProspect
        saved_rows = db.query(SavedProspect).filter(SavedProspect.player_id == player.id).all()
        player_name = f"{player.first_name} {player.last_name}"
        for sp in saved_rows:
            create_notification(
                db,
                sp.scout_id,
                "star",
                "Player Status Update",
                f"{player_name}'s status changed to {body.status}.",
            )

    db.commit()
    db.refresh(player)

    return _player_item(player, club.name if club else None)


@router.patch("/players/{player_id}/stats", response_model=PlayerStats)
def update_player_stats(
    player_id: str,
    body: UpdatePlayerStatsRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    player = _get_player_or_404(db, player_id)

    player.minutes_played = body.minutes_played
    player.goals = body.goals
    player.assists = body.assists
    player.saves = body.saves
    player.defensive_contributions = body.defensive_contributions
    player.chances_created = body.chances_created
    player.dribbles = body.dribbles

    db.commit()
    db.refresh(player)

    return PlayerStats.from_player(player)


@router.delete("/players/{player_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_player(
    player_id: str,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
) -> None:
    player = _get_player_or_404(db, player_id)
    db.delete(player)
    db.commit()


# ---------------------------------------------------------------------------
# Reports – Bulk delete
# ---------------------------------------------------------------------------

@router.post("/reports/bulk-delete", response_model=BulkDeleteResult)
def bulk_delete_reports(
    body: BulkDeleteRequest,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    uuids = parse_uuid_list(body.ids)
    if not uuids:
        return BulkDeleteResult(deleted=0)
    count = db.query(ScoutingReport).filter(ScoutingReport.id.in_(uuids)).delete(synchronize_session=False)
    db.commit()
    return BulkDeleteResult(deleted=count)


# ---------------------------------------------------------------------------
# Reports – Update & Delete
# ---------------------------------------------------------------------------

@router.put("/reports/{report_id}", response_model=AdminReportItem)
def update_report(
    report_id: str,
    body: UpdateReportRequest,
    background_tasks: BackgroundTasks,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
):
    report = _get_report_or_404(db, report_id)

    if body.player_id:
        report.player_id = _get_player_or_404(db, body.player_id).id

    old_status = report.status
    report.player_name = body.player_name.strip()
    report.position = body.position.strip()
    report.rating = body.rating
    report.status = body.status
    report.notes = body.notes
    report.updated_at = datetime.now(timezone.utc)

    if body.status in ("approved", "rejected") and body.status != old_status:
        create_notification(
            db,
            report.scout_id,
            "file",
            f"Report {body.status.capitalize()}",
            f"Your report for {body.player_name.strip()} has been {body.status}.",
        )

    scout = db.query(User).filter(User.id == report.scout_id).first()
    db.commit()
    db.refresh(report)

    if body.status in ("approved", "rejected") and body.status != old_status:
        if scout and scout.telegram_chat_id:
            background_tasks.add_task(
                send_report_notification,
                scout.telegram_chat_id,
                report.player_name,
                body.status,
            )

    return _report_item(report, f"{scout.first_name} {scout.last_name}" if scout else "Unknown")


@router.delete("/reports/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_report(
    report_id: str,
    _: User = Depends(require_global_admin),
    db: Session = Depends(get_db),
) -> None:
    report = _get_report_or_404(db, report_id)
    db.delete(report)
    db.commit()
